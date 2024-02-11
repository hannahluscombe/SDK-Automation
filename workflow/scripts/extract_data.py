"""Module to extract txt file data from xlsx SDK

This is the python equivalent to the VBA macro in the starter datakit files.
Note, takes about 15-20sec per file
"""

from openpyxl import load_workbook
import sys

def excel_to_text(xlsm: str, out_file: str):
    wb = load_workbook(xlsm, data_only=True, keep_links=False, read_only=True)
    ws = wb["ToDataFile"]

    with open(out_file, "w") as f:
        for row in ws.iter_rows(values_only=True):
            data = " ".join(str(cell) for cell in row if cell is not None)
            f.write(f"{data}\n")

if __name__ == "__main__":
    
    if "snakemake" in globals():
        excel_to_text(snakemake.input.xlsm, snakemake.output.txt)
    else:
        if len(sys.argv) != 3:
            msg = "Usage: python {} <xlsm_file_in> <txt_file_out>"
            print(msg.format(sys.argv[0]))
            sys.exit(1)
        else:
            xlsm = sys.argv[1]
            txt_file = sys.argv[2]
            excel_to_text(xlsm, txt_file)